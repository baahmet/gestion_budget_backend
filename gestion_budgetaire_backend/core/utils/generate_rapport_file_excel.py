import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.chart import PieChart, Reference
from django.core.files.base import ContentFile
from django.utils.timezone import now


def generate_rapport_file_excel(budget, recettes, depenses, commandes, periode, user):
    wb = Workbook()
    ws = wb.active
    ws.title = "Rapport Financier"

    # ========== Styles ==========
    title_font = Font(bold=True, size=16)
    header_font = Font(bold=True, color="FFFFFF")
    section_font = Font(bold=True)
    italic_font = Font(italic=True)
    center_align = Alignment(horizontal="center")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

    row = 1

    # ========== Titre ==========
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    ws.cell(row=row, column=1, value=f"RAPPORT FINANCIER - {budget.exercice}")
    ws.cell(row=row, column=1).font = title_font
    ws.cell(row=row, column=1).alignment = center_align
    row += 2

    # ========== Infos générales ==========
    ws.cell(row=row, column=1, value="Période :")
    ws.cell(row=row, column=2, value=periode)
    row += 1
    ws.cell(row=row, column=1, value="Montant Total :")
    ws.cell(row=row, column=2, value=f"{budget.montant_total:,.0f} FCFA")
    row += 1
    ws.cell(row=row, column=1, value="Montant Disponible :")
    ws.cell(row=row, column=2, value=f"{budget.montant_disponible:,.0f} FCFA")
    row += 2

    # ========== Recettes ==========
    ws.cell(row=row, column=1, value="Recettes").font = section_font
    row += 1
    headers = ["Source", "Type", "Montant"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = center_align
    row += 1
    for r in recettes:
        ws.cell(row=row, column=1, value=r.source).border = thin_border
        ws.cell(row=row, column=2, value=r.type).border = thin_border
        ws.cell(row=row, column=3, value=r.montant).border = thin_border
        row += 1
    row += 2

    # ========== Dépenses ==========
    ws.cell(row=row, column=1, value="Dépenses validées").font = section_font
    row += 1
    headers = ["Type", "Catégorie", "Montant"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = center_align
    row += 1
    for d in depenses:
        ws.cell(row=row, column=1, value=d.type_depense).border = thin_border
        ws.cell(row=row, column=2, value=d.categorie).border = thin_border
        ws.cell(row=row, column=3, value=d.montant).border = thin_border
        row += 1
    row += 2

    # ========== Commandes ==========
    ws.cell(row=row, column=1, value="Commandes").font = section_font
    row += 1
    headers = ["Désignation", "Quantité", "Total"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = center_align
    row += 1
    for c in commandes:
        ws.cell(row=row, column=1, value=c.designation).border = thin_border
        ws.cell(row=row, column=2, value=c.quantite).border = thin_border
        ws.cell(row=row, column=3, value=c.total).border = thin_border
        row += 1
    row += 2

    # ========== Pied de page ==========
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    ws.cell(row=row, column=1, value=f"Généré par : {user.nom} ({user.role}) - {now().strftime('%d/%m/%Y %H:%M')}")
    ws.cell(row=row, column=1).font = italic_font
    row += 2

    # ========== Données pour le graphique ==========
    ws.cell(row=row, column=1, value="📊 Répartition du budget").font = section_font
    row += 1
    ws.cell(row=row, column=1, value="Montant Disponible")
    ws.cell(row=row + 1, column=1, value="Dépenses Validées")
    ws.cell(row=row, column=2, value=budget.montant_disponible)
    ws.cell(row=row + 1, column=2, value=budget.montant_total_depenses_validees)

    # Graphique circulaire
    pie = PieChart()
    pie.title = "Répartition : Disponible vs Dépensé"
    labels = Reference(ws, min_col=1, min_row=row, max_row=row + 1)
    data = Reference(ws, min_col=2, min_row=row, max_row=row + 1)
    pie.add_data(data, titles_from_data=False)
    pie.set_categories(labels)
    pie.height = 7
    pie.width = 9
    ws.add_chart(pie, f"D{row}")

    # ========== Ajustement de colonnes ==========
    for col in range(1, 6):
        ws.column_dimensions[chr(64 + col)].width = 25

    # ========== Sauvegarde ==========
    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)

    filename = f"rapport_{budget.exercice}_{periode}.xlsx".replace(" ", "_")
    django_file = ContentFile(file_stream.read(), name=filename)
    return django_file, filename
